import dataclasses
import uuid
import inspect
import openpyxl
import pandas as pd
import openpyxl as opyxl

import ifrs17.DataStructure
from .Systemorph import *

_formats: dict[str, Callable] = {}


@dataclass
class Options:
    TargetDataSource: IDataSource


def _FromFileWithFormat(pathToFile: str, target: IDataSource, format_: str):
    _formats[format_](Options(target), _FromFileToDataSet(pathToFile))


def DefineFormat(format: str, body: Callable[Options, IDataSet]):
    _formats[format] = body


def _FromFileToDataSet(pathToFile: str):
    result: dict[str, pd.DataFrame] = {}
    for name in opyxl.load_workbook(pathToFile).sheetnames:
        result[name] = pd.read_excel(pathToFile, sheet_name=name, na_values=[''])

    return IDataSet(result)


def df_to_records(df: pd.DataFrame, type_: type) -> list:

    # fields = [v for k, v in AocTypes.__dict__.items() if k[0] != '_']
    records = []
    try:
        for r in df.to_dict('records'):
            records.append(type_(**r))

    except TypeError:
        records.clear()
        for r in df.to_dict('records'):
            r['Id'] = uuid.uuid4()
            records.append(type_(**r))

    return records


def FromDataSet(dataSet: IDataSet, type_: type, workspace: IWorkspace, body: Callable=None, format_=None):

    if body:
        if format_:
            name = format_
        else:
            name = type_.__name__
        result = []
        for r in dataSet.Tables[name].to_dict('records'):
            result.append(body(dataSet, r))
        workspace.UpdateAsync(type_, result)
    else:
        workspace.UpdateAsync(type_, df_to_records(dataSet.Tables[type_.__name__], type_))


def FromFile(pathToFile: str, target: IDataSource, format_: str = '',
             type_: Optional[Union[type, list[type]]] = None):
    if type_:
        result = _FromFileWithType(pathToFile, type_)
        for k, v in result.items():
            target.UpdateAsync(k, v)
    else:
        _FromFileWithFormat(pathToFile, target, format_)


def _FromFileWithType(pathToFile: str,
             type_: Optional[Union[type, list[type]]] = None):

    types_ = type_ if isinstance(type_, Iterable) else [type_]

    wb = openpyxl.load_workbook(pathToFile)
    result = {}
    for tp in types_:
        rows = list(wb[tp.__name__].values)
        objs = []
        for r in rows[1:]:
            kwargs = {}
            extids = {}     # index: value
            vals = {}       # index: value
            for i, name in enumerate(rows[0]):
                if name[:10] == "ExternalId":
                    extids[int(name[10:])] = r[i] if r[i] else ''
                elif name[:6] == "Values":
                    vals[int(name[6:])] = float(r[i])
                elif name == "InputSource":
                    kwargs[name] = [s.strip() for s in r[i].split(",")]
                else:
                    kwargs[name] = r[i]

            if "ExternalId" in inspect.signature(tp).parameters.keys():
                kwargs["ExternalId"] = list(extids[i] for i in extids.keys())
            elif "Values" in inspect.signature(tp).parameters.keys():
                kwargs["Values"] = list(vals[i] for i in vals.keys())

            fields = set(f.name for f in dataclasses.fields(tp))
            args = set(kwargs)
            missing = fields - args

            if missing:
                if 'Id' in missing:
                    kwargs['Id'] = uuid.uuid4()
                if 'Name' in missing:
                    kwargs['Name'] = ''
                if 'Scenario' in missing:
                    kwargs['Scenario'] = ''

            objs.append(tp(**kwargs))


        result[tp] = objs

    return result


